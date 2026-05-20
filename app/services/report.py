from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def build_excel_report(payload: dict[str, Any]) -> bytes:
    rows = payload.get("port_checks", {}).get("results", [])
    if not rows:
        raise ValueError("No check results available.")

    flattened = []
    checked_at = payload.get("checked_at")
    for row in rows:
        flattened.append(
            {
                "Checked At (UTC)": checked_at,
                "Server Name": row.get("server_name"),
                "Host": row.get("host"),
                "Port": row.get("port"),
                "Status": row.get("status"),
                "Detail": row.get("detail"),
                "Latency (ms)": row.get("latency_ms"),
            }
        )

    df = pd.DataFrame(flattened)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheet_name = "Infra Report"
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        thin = Side(style="thin", color="9AA4B2")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="E0F2FE")
        header_font = Font(bold=True, color="0F172A")

        status_fills = {
            "OPEN": PatternFill("solid", fgColor="DCFCE7"),
            "REFUSED": PatternFill("solid", fgColor="FEE2E2"),
            "TIMEOUT": PatternFill("solid", fgColor="FEE2E2"),
            "UNKNOWN_HOST": PatternFill("solid", fgColor="FDE68A"),
            "ERROR": PatternFill("solid", fgColor="FECACA"),
        }

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                if cell.column == 5 and cell.value in status_fills:
                    cell.fill = status_fills[cell.value]

        widths = {
            "A": 24,
            "B": 20,
            "C": 18,
            "D": 10,
            "E": 14,
            "F": 50,
            "G": 12,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        ws.auto_filter.ref = ws.dimensions

    output.seek(0)
    return output.getvalue()

